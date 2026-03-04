"""myuni — Grade import tool for Greek university XLSX gradebooks.

Usage:
    uv run main.py <csv_file> <xlsx_file> [--only-passing] [--debug]

See README.md for full documentation.
"""

import argparse
import openpyxl

# Column headers as they appear in the university-issued XLSX export.
# These are Greek-language strings and must not be changed.
HEADER_STUDENT_ID = "Αριθμός Μητρώου"
HEADER_GRADE = "Βαθμολογία"

# Institutional constants derived from the university's student ID scheme.
STUDENT_ID_PREFIX = "111520"
STUDENT_ID_LENGTH = 13
STUDENT_ID_FIXED_SEGMENT_SLICE = slice(8, 10)
STUDENT_ID_FIXED_SEGMENT_VALUE = "00"

MAX_GRADE = 10
PASSING_GRADE = 5


def read_grades_from_csv(file_path: str) -> dict[str, int]:
    """Read and validate student grades from a CSV file.

    Each non-empty line must follow the format::

        STUDENT_ID,GRADE

    where STUDENT_ID is a 13-digit numeric string of the form::

        111520YY00NNN
        ^^^^^^ ^^ ^^ ^^^
        |      |  |  └── student number (3 digits)
        |      |  └───── fixed segment "00"
        |      └──────── matriculation year (2 digits)
        └─────────────── fixed institutional prefix

    and GRADE is an integer.

    Args:
        file_path: Path to the CSV file.

    Returns:
        A dict mapping each student ID string to its integer grade.

    Raises:
        ValueError: If any line has an invalid format, an unrecognised student
            ID, or a non-integer grade value.
    """
    grades: dict[str, int] = {}

    with open(file_path, "r") as fh:
        lines = fh.readlines()

    for line in lines:
        line = line.strip()
        if not line:
            continue

        parts = line.split(",")
        if len(parts) != 2:
            raise ValueError(f"Invalid line format: {line!r}")

        student_id, grade_str = parts

        # Validate student ID structure.
        if not student_id.startswith(STUDENT_ID_PREFIX):
            raise ValueError(f"Invalid student ID prefix: {student_id!r}")
        if len(student_id) < STUDENT_ID_LENGTH:
            raise ValueError(f"Student ID too short: {student_id!r}")
        if not student_id.isdigit():
            raise ValueError(f"Student ID must be numeric: {student_id!r}")
        if student_id[STUDENT_ID_FIXED_SEGMENT_SLICE] != STUDENT_ID_FIXED_SEGMENT_VALUE:
            raise ValueError(
                f"Invalid student ID format "
                f"(expected {STUDENT_ID_FIXED_SEGMENT_VALUE!r} at positions 8-9): "
                f"{student_id!r}"
            )

        try:
            grade = int(grade_str)
        except ValueError:
            raise ValueError(f"Invalid grade for student {student_id!r}: {grade_str!r}")

        grades[student_id] = grade

    return grades


def apply_grades_to_xlsx(
    grades: dict[str, int],
    xlsx_file: str,
    debug: bool = False,
    unregistered: bool = False,
) -> None:
    """Write grades from *grades* into the matching rows of an XLSX workbook.

    The active sheet is expected to have:
    - Row 1: a title/meta row (skipped).
    - Row 2: column headers, including ``HEADER_STUDENT_ID`` and ``HEADER_GRADE``.
    - Row 3+: one student per row.

    For each data row, the student ID cell is compared against the keys of
    *grades*.  When a match is found the corresponding grade is written into
    the grade cell.  Grades above ``MAX_GRADE`` are silently capped.

    Args:
        grades: Mapping of student ID strings to integer grades.
        xlsx_file: Path to the XLSX file to update (written in-place).
        debug: When ``True``, log a message for every student row in the
            workbook whose ID is not present in *grades*.
        unregistered: When ``True``, print every student ID present in
            *grades* (CSV) but absent from the XLSX (took the exam but
            never registered).

    Raises:
        AssertionError: If the expected column headers are not found in row 2.
    """
    workbook = openpyxl.load_workbook(xlsx_file)
    sheet = workbook.active

    # Locate the header row (row 2) to find column indices dynamically.
    # This makes the tool resilient to column reordering.
    student_id_index: int | None = None
    grade_index: int | None = None
    applied_grades = 0
    total_students = 0

    # iter_rows starts at min_row=2; the first row yielded is the header row.
    parsing_headers = True
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if parsing_headers:
            parsing_headers = False
            for cell in row:
                if cell.value == HEADER_STUDENT_ID:
                    student_id_index = cell.column - 1  # convert to 0-based index
                elif cell.value == HEADER_GRADE:
                    grade_index = cell.column - 1
            continue

        assert student_id_index is not None, (
            f"Student ID column {HEADER_STUDENT_ID!r} not found in header row."
        )
        assert grade_index is not None, (
            f"Grade column {HEADER_GRADE!r} not found in header row."
        )

        total_students += 1
        student_id_cell = row[student_id_index]
        grade_cell = row[grade_index]
        student_id = str(student_id_cell.value)

        if student_id in grades:
            grade = grades[student_id]
            if grade > MAX_GRADE:
                grade = MAX_GRADE
                print(f"Grade for student {student_id} capped at {MAX_GRADE}.")
            grade_cell.value = grade
            print(f"Applied grade {grade} to student {student_id}.")
            applied_grades += 1
        else:
            if debug:
                print(f"Student {student_id} not found in CSV — skipping.")

    print(f"Total students processed: {total_students}, Grades applied: {applied_grades}")

    if unregistered and student_id_index is not None:
        xlsx_ids = {
            str(row[student_id_index].value)
            for row in sheet.iter_rows(min_row=3, values_only=False)
            if row[student_id_index].value is not None
        }
        missing = sorted(sid for sid in grades if sid not in xlsx_ids)
        if missing:
            print(f"\nUnregistered students ({len(missing)} took exam but not in XLSX):")
            for sid in missing:
                print(f"  {sid} (grade: {grades[sid]})")
        else:
            print("\nNo unregistered students found.")

    workbook.save(xlsx_file)
    print(f"Grades saved to {xlsx_file}.")


def main() -> None:
    """Parse CLI arguments and run the grade import pipeline."""
    parser = argparse.ArgumentParser(
        description=(
            "Import student grades from a CSV file into a university XLSX gradebook. "
            "Grades are matched by student ID and written into the grade column."
        )
    )
    parser.add_argument(
        "csv_file",
        help="Path to the CSV file containing grades (format: STUDENT_ID,GRADE per line).",
    )
    parser.add_argument(
        "xlsx_file",
        help="Path to the XLSX gradebook where grades will be written.",
    )
    parser.add_argument(
        "--only-passing",
        action="store_true",
        help=f"Only apply grades >= {PASSING_GRADE} (passing threshold). "
             "Non-passing grades are excluded from the import.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print a message for each student in the XLSX whose ID is not in the CSV.",
    )
    parser.add_argument(
        "--unregistered",
        action="store_true",
        help="Print all students in the CSV (took the exam) who are absent from the XLSX "
             "(i.e., never registered for the course).",
    )
    args = parser.parse_args()

    grades = read_grades_from_csv(args.csv_file)

    if args.only_passing:
        grades = {sid: g for sid, g in grades.items() if g >= PASSING_GRADE}

    apply_grades_to_xlsx(grades, args.xlsx_file, debug=args.debug, unregistered=args.unregistered)


if __name__ == "__main__":
    main()
